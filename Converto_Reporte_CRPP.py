import customtkinter
from CTkMessagebox import CTkMessagebox
import tkinter
import os
from PIL import Image
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
from pathlib import Path
from tkinter import filedialog

def Corto_Plazo_Mun(directorio_in, directorio_out):

    def ejecutar_script_1():
        # Definir el ancho fijo de las columnas crpp4601
        column_widths = [28, 9, 12, 4, 4, 9, 13, 6, 11]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_VENCIMIENTO", "CUOTAS", "SALDO" 
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4601 o crpp-4601
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4601' not in contenido and 'crpp-4601' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4601 o crpp-4601, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('MUN', case=False, na=False)]
                    
                    # Verificar que no repita la "F.VENCIMIENTO"
                    fechas_en_archivo = set(filtered_data['F_VENCIMIENTO'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["F_VENCIMIENTO"] = pd.to_datetime(filtered_data["F_VENCIMIENTO"], format="%d/%m/%Y")
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "CUOTAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 1.")
            return pd.DataFrame(columns=column_names)

    def ejecutar_script_2():
        # Definir el ancho fijo de las columnas crpp4610
        column_widths = [28, 9, 12, 4, 4, 8, 12, 4, 15, 14, 13, 14]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_AMORTIZACION", "DIAS", "SALDO", 
            "PROV.ANT", "PROV.MES", "PROV.ACTUAL"
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4610 o crpp-4610
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4610' not in contenido and 'crpp-4610' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4610 o crpp-4610, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('MUN', case=False, na=False)]
                    
                    # Verificar que no repita la "F_AMORTIZACION"
                    fechas_en_archivo = set(filtered_data['F_AMORTIZACION'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las fechas, manejando fechas inválidas
                    def parse_fecha(fecha):
                        try:
                            return datetime.datetime.strptime(fecha, "%d/%m/%Y").strftime("%d/%m/%Y")
                        except ValueError:
                            return "00/00/0000"
                    
                    filtered_data["F_AMORTIZACION"] = filtered_data["F_AMORTIZACION"].apply(parse_fecha)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ANT"] = filtered_data["PROV.ANT"].str.replace(',', '').astype(float)
                    filtered_data["PROV.MES"] = filtered_data["PROV.MES"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ACTUAL"] = filtered_data["PROV.ACTUAL"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "DIAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 2.")
            return pd.DataFrame(columns=column_names)

    # Ejecutar ambos scripts y guardar los resultados en variables
    df_script_1 = ejecutar_script_1()
    df_script_2 = ejecutar_script_2()

    df_script_1["PK1"] = df_script_1["N°CUENTA"] + df_script_1["PAGARE"]
    df_script_2["PK1"] = df_script_2["N°CUENTA"] + df_script_2["PAGARE"]

    # Realizar el join por "PK1"
    resultado_final = pd.merge(df_script_1, df_script_2, on="PK1", how="inner", suffixes=('_script1', '_script2'))

    # Crear un nuevo Workbook con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    df= resultado_final.drop([resultado_final.columns[9], resultado_final.columns[10] ,resultado_final.columns[11], resultado_final.columns[12], resultado_final.columns[13]], axis=1)  

    # Escribir el DataFrame al Worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es la primera fila (encabezados)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if 'F_VENCIMIENTO' in df.columns and c_idx == df.columns.get_loc('F_VENCIMIENTO') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif 'F_AMORTIZACION' in df.columns and c_idx == df.columns.get_loc('F_AMORTIZACION') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif c_idx in [df.columns.get_loc('TASA_script1') + 1, 
                            df.columns.get_loc('SALDO_script1') + 1,
                            df.columns.get_loc('PROV.ANT') + 1,
                            df.columns.get_loc('PROV.MES') + 1,
                            df.columns.get_loc('PROV.ACTUAL') + 1]:
                    cell.number_format = '#,##0.00'

    # Ajustar el ancho de las columnas
    column_widths_adjusted = [30, 11, 15, 10, 10, 10, 15, 10, 20, 30, 11, 15, 10, 10, 10, 12, 10, 15, 14, 13, 20]
    for idx, width in enumerate(column_widths_adjusted, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Guardar el Workbook
    output_file = os.path.join(directorio_out, "Reporte-Corto-Plazo-MUN-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    wb.save(output_file)

    print(f"El archivo consolidado se ha guardado en: {output_file}")

def Corto_plazo_No_Mun():
    directorio_in = "D:/CRPP/in"
    directorio_out = "D:/CRPP/out"

    def ejecutar_script_1():
        # Definir el ancho fijo de las columnas crpp4601
        column_widths = [28, 9, 12, 4, 4, 9, 13, 6, 11]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_VENCIMIENTO", "CUOTAS", "SALDO" 
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4601 o crpp-4601
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4601' not in contenido and 'crpp-4601' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4601 o crpp-4601, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('ELEC|EMP', case=False, na=False)]
                    
                    # Verificar que no repita la "F.VENCIMIENTO"
                    fechas_en_archivo = set(filtered_data['F_VENCIMIENTO'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["F_VENCIMIENTO"] = pd.to_datetime(filtered_data["F_VENCIMIENTO"], format="%d/%m/%Y")
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "CUOTAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 1.")
            return pd.DataFrame(columns=column_names)

    def ejecutar_script_2():
        # Definir el ancho fijo de las columnas crpp4610
        column_widths = [28, 9, 12, 4, 4, 8, 12, 4, 15, 14, 13, 14]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_AMORTIZACION", "DIAS", "SALDO", 
            "PROV.ANT", "PROV.MES", "PROV.ACTUAL"
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4610 o crpp-4610
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4610' not in contenido and 'crpp-4610' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4610 o crpp-4610, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('ELEC|EMP', case=False, na=False)]
                    
                    # Verificar que no repita la "F_AMORTIZACION"
                    fechas_en_archivo = set(filtered_data['F_AMORTIZACION'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las fechas, manejando fechas inválidas
                    def parse_fecha(fecha):
                        try:
                            return datetime.datetime.strptime(fecha, "%d/%m/%Y").strftime("%d/%m/%Y")
                        except ValueError:
                            return "00/00/0000"
                    
                    filtered_data["F_AMORTIZACION"] = filtered_data["F_AMORTIZACION"].apply(parse_fecha)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ANT"] = filtered_data["PROV.ANT"].str.replace(',', '').astype(float)
                    filtered_data["PROV.MES"] = filtered_data["PROV.MES"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ACTUAL"] = filtered_data["PROV.ACTUAL"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "DIAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 2.")
            return pd.DataFrame(columns=column_names)

    # Ejecutar ambos scripts y guardar los resultados en variables
    df_script_1 = ejecutar_script_1()
    df_script_2 = ejecutar_script_2()

    df_script_1["PK1"] = df_script_1["N°CUENTA"] + df_script_1["PAGARE"]
    df_script_2["PK1"] = df_script_2["N°CUENTA"] + df_script_2["PAGARE"]

    # Realizar el join por "N°CUENTA"
    resultado_final = pd.merge(df_script_1, df_script_2, on="PK1", how="inner", suffixes=('_script1', '_script2'))

    # Crear un nuevo Workbook con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    df= resultado_final.drop([resultado_final.columns[9], resultado_final.columns[10] ,resultado_final.columns[11], resultado_final.columns[12], resultado_final.columns[13]], axis=1)  

    # Escribir el DataFrame al Worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es la primera fila (encabezados)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if 'F_VENCIMIENTO' in df.columns and c_idx == df.columns.get_loc('F_VENCIMIENTO') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif 'F_AMORTIZACION' in df.columns and c_idx == df.columns.get_loc('F_AMORTIZACION') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif c_idx in [df.columns.get_loc('TASA_script1') + 1, 
                            df.columns.get_loc('SALDO_script1') + 1,
                            df.columns.get_loc('PROV.ANT') + 1,
                            df.columns.get_loc('PROV.MES') + 1,
                            df.columns.get_loc('PROV.ACTUAL') + 1]:
                    cell.number_format = '#,##0.00'

    # Ajustar el ancho de las columnas
    column_widths_adjusted = [30, 11, 15, 10, 10, 10, 15, 10, 20, 30, 11, 15, 10, 10, 10, 12, 10, 15, 14, 13, 14]
    for idx, width in enumerate(column_widths_adjusted, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Guardar el Workbook
    output_file = os.path.join(directorio_out, "Reporte-Corto-Plazo-NO-MUN-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    wb.save(output_file)

    print(f"El archivo consolidado se ha guardado en: {output_file}")

def Largo_Plazo_Mun():
    # Configuración de directorios
    directorio_in = "D:/CRPP/in"
    directorio_out = "D:/CRPP/out"

    def ejecutar_script_1():
        # Definir el ancho fijo de las columnas crpp4601
        column_widths = [28, 9, 12, 4, 4, 9, 13, 6, 11]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_VENCIMIENTO", "CUOTAS", "SALDO" 
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4601 o crpp-4601
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4602' not in contenido and 'crpp-4602' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4602 o crpp-4602, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('MUN', case=False, na=False)]
                    
                    # Verificar que no repita la "F.VENCIMIENTO"
                    fechas_en_archivo = set(filtered_data['F_VENCIMIENTO'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["F_VENCIMIENTO"] = pd.to_datetime(filtered_data["F_VENCIMIENTO"], format="%d/%m/%Y")
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "CUOTAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 1.")
            return pd.DataFrame(columns=column_names)

    def ejecutar_script_2():
        # Definir el ancho fijo de las columnas crpp4610
        column_widths = [28, 9, 12, 4, 4, 8, 12, 4, 15, 14, 13, 14]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_AMORTIZACION", "DIAS", "SALDO", 
            "PROV.ANT", "PROV.MES", "PROV.ACTUAL"
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4610 o crpp-4610
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4628' not in contenido and 'crpp-4628' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4610 o crpp-4610, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('MUN', case=False, na=False)]
                    
                    # Verificar que no repita la "F_AMORTIZACION"
                    fechas_en_archivo = set(filtered_data['F_AMORTIZACION'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las fechas, manejando fechas inválidas
                    def parse_fecha(fecha):
                        try:
                            return datetime.datetime.strptime(fecha, "%d/%m/%Y").strftime("%d/%m/%Y")
                        except ValueError:
                            return "00/00/0000"
                    
                    filtered_data["F_AMORTIZACION"] = filtered_data["F_AMORTIZACION"].apply(parse_fecha)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ANT"] = filtered_data["PROV.ANT"].str.replace(',', '').astype(float)
                    filtered_data["PROV.MES"] = filtered_data["PROV.MES"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ACTUAL"] = filtered_data["PROV.ACTUAL"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "DIAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 2.")
            return pd.DataFrame(columns=column_names)

    # Ejecutar ambos scripts y guardar los resultados en variables
    df_script_1 = ejecutar_script_1()
    df_script_2 = ejecutar_script_2()

    df_script_1["PK1"] = df_script_1["N°CUENTA"] + df_script_1["PAGARE"]
    df_script_2["PK1"] = df_script_2["N°CUENTA"] + df_script_2["PAGARE"]

    # Realizar el join por "PK1"
    resultado_final = pd.merge(df_script_1, df_script_2, on="PK1", how="inner", suffixes=('_script1', '_script2'))

    # Crear un nuevo Workbook con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    print(resultado_final)

    #re= pd.DataFrame(resultado_final[1:], columns=resultado_final[0])
    df= resultado_final.drop([resultado_final.columns[9], resultado_final.columns[10] ,resultado_final.columns[11], resultado_final.columns[12], resultado_final.columns[13]], axis=1)  

    # Escribir el DataFrame al Worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es la primera fila (encabezados)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if 'F_VENCIMIENTO' in df.columns and c_idx == df.columns.get_loc('F_VENCIMIENTO') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif 'F_AMORTIZACION' in df.columns and c_idx == df.columns.get_loc('F_AMORTIZACION') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif c_idx in [df.columns.get_loc('TASA_script1') + 1, 
                            df.columns.get_loc('SALDO_script1') + 1,
                            df.columns.get_loc('PROV.ANT') + 1,
                            df.columns.get_loc('PROV.MES') + 1,
                            df.columns.get_loc('PROV.ACTUAL') + 1]:
                    cell.number_format = '#,##0.00'


    # Ajustar el ancho de las columnas
    column_widths_adjusted = [30, 10, 15, 10, 10, 10, 15, 10, 20, 30, 11, 15, 10, 10, 10, 12, 10, 15, 14, 13, 20]
    for idx, width in enumerate(column_widths_adjusted, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Guardar el Workbook
    output_file = os.path.join(directorio_out, "Reporte-Largo-Plazo-MUN-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    wb.save(output_file)

    print(f"El archivo consolidado se ha guardado en: {output_file}")

def Largo_Plazo_No_Mun():
    # Configuración de directorios
    directorio_in = "D:/CRPP/in"
    directorio_out = "D:/CRPP/out"

    def ejecutar_script_1():
        # Definir el ancho fijo de las columnas crpp4601
        column_widths = [28, 9, 12, 4, 4, 9, 13, 6, 11]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_VENCIMIENTO", "CUOTAS", "SALDO" 
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4601 o crpp-4601
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4602' not in contenido and 'crpp-4602' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4602 o crpp-4602, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('ELEC|EMP', case=False, na=False)]
                    
                    # Verificar que no repita la "F.VENCIMIENTO"
                    fechas_en_archivo = set(filtered_data['F_VENCIMIENTO'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["F_VENCIMIENTO"] = pd.to_datetime(filtered_data["F_VENCIMIENTO"], format="%d/%m/%Y")
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "CUOTAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 1.")
            return pd.DataFrame(columns=column_names)

    def ejecutar_script_2():
        # Definir el ancho fijo de las columnas crpp4610
        column_widths = [28, 9, 12, 4, 4, 8, 12, 4, 15, 14, 13, 14]
        column_names = [
            "CLIENTE", "CODIGO", "N°CUENTA", "PAGARE", "FE", "TASA", "F_AMORTIZACION", "DIAS", "SALDO", 
            "PROV.ANT", "PROV.MES", "PROV.ACTUAL"
        ]

        # Lista para almacenar los DataFrames consolidados
        consolidated_data = []

        # Set para almacenar las fechas de proceso encontradas
        fechas_proceso = set()

        # Obtén la lista de archivos en el directorio
        archivos = os.listdir(directorio_in)

        for archivo in archivos:
            if archivo.lower().endswith(".txt"):
                ruta_completa = os.path.join(directorio_in, archivo)
                
                # Validar que el archivo contenga CRPP-4610 o crpp-4610
                with open(ruta_completa, 'r') as file:
                    contenido = file.read()
                    if 'CRPP-4628' not in contenido and 'crpp-4628' not in contenido:
                        print(f"El archivo {archivo} no contiene CRPP-4610 o crpp-4610, será ignorado.")
                        continue
                
                # Intenta leer el archivo de texto en un DataFrame
                try:
                    data = pd.read_fwf(ruta_completa, widths=column_widths, skiprows=8, names=column_names)

                    # Filtrar por el valor específico en la primera columna
                    filtered_data = data[data['CLIENTE'].str.contains('ELEC|EMP', case=False, na=False)]
                    
                    # Verificar que no repita la "F_AMORTIZACION"
                    fechas_en_archivo = set(filtered_data['F_AMORTIZACION'])
                    if fechas_en_archivo & fechas_proceso:
                        print(f"El archivo {archivo} contiene fechas de procesamiento ya procesadas, será rechazado.")
                        continue
                    fechas_proceso.update(fechas_en_archivo)
                    
                    # Convertir las fechas, manejando fechas inválidas
                    def parse_fecha(fecha):
                        try:
                            return datetime.datetime.strptime(fecha, "%d/%m/%Y").strftime("%d/%m/%Y")
                        except ValueError:
                            return "00/00/0000"
                    
                    filtered_data["F_AMORTIZACION"] = filtered_data["F_AMORTIZACION"].apply(parse_fecha)
                    
                    # Convertir las columnas al tipo de dato adecuado
                    filtered_data["TASA"] = filtered_data["TASA"].str.replace(',', '').astype(float)
                    filtered_data["SALDO"] = filtered_data["SALDO"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ANT"] = filtered_data["PROV.ANT"].str.replace(',', '').astype(float)
                    filtered_data["PROV.MES"] = filtered_data["PROV.MES"].str.replace(',', '').astype(float)
                    filtered_data["PROV.ACTUAL"] = filtered_data["PROV.ACTUAL"].str.replace(',', '').astype(float)
                    
                    # Agregar los datos filtrados a la lista de datos consolidados
                    consolidated_data.append(filtered_data)

                except Exception as e:
                    print(f"No se pudo leer el archivo {archivo}: {e}")

        if consolidated_data:
            resultado_df = pd.concat(consolidated_data, ignore_index=True)
            resultado_df = resultado_df.astype({
                "CLIENTE": str,
                "CODIGO": str,
                "N°CUENTA": str,
                "PAGARE": str,
                "FE": str,
                "DIAS": str,
            })
            return resultado_df
        else:
            print("No se encontraron datos para consolidar en Script 2.")
            return pd.DataFrame(columns=column_names)

    # Ejecutar ambos scripts y guardar los resultados en variables
    df_script_1 = ejecutar_script_1()
    df_script_2 = ejecutar_script_2()

    df_script_1["PK1"] = df_script_1["N°CUENTA"] + df_script_1["PAGARE"]
    df_script_2["PK1"] = df_script_2["N°CUENTA"] + df_script_2["PAGARE"]

    # Realizar el join por "PK1"
    resultado_final = pd.merge(df_script_1, df_script_2, on="PK1", how="inner", suffixes=('_script1', '_script2'))

    # Crear un nuevo Workbook con openpyxl
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    print(resultado_final)

    #re= pd.DataFrame(resultado_final[1:], columns=resultado_final[0])
    df= resultado_final.drop([resultado_final.columns[9], resultado_final.columns[10] ,resultado_final.columns[11], resultado_final.columns[12], resultado_final.columns[13]], axis=1)  

    # Escribir el DataFrame al Worksheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:  # Si es la primera fila (encabezados)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                if 'F_VENCIMIENTO' in df.columns and c_idx == df.columns.get_loc('F_VENCIMIENTO') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif 'F_AMORTIZACION' in df.columns and c_idx == df.columns.get_loc('F_AMORTIZACION') + 1:
                    cell.number_format = 'DD/MM/YYYY'
                elif c_idx in [df.columns.get_loc('TASA_script1') + 1, 
                            df.columns.get_loc('SALDO_script1') + 1,
                            df.columns.get_loc('PROV.ANT') + 1,
                            df.columns.get_loc('PROV.MES') + 1,
                            df.columns.get_loc('PROV.ACTUAL') + 1]:
                    cell.number_format = '#,##0.00'


    # Ajustar el ancho de las columnas
    column_widths_adjusted = [30, 10, 15, 10, 10, 10, 15, 10, 20, 30, 11, 15, 10, 10, 10, 12, 10, 15, 14, 13, 14]
    for idx, width in enumerate(column_widths_adjusted, 1):
        ws.column_dimensions[chr(64 + idx)].width = width

    # Guardar el Workbook
    output_file = os.path.join(directorio_out, "Reporte-Largo-Plazo-NO-MUN-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    wb.save(output_file)

    print(f"El archivo consolidado se ha guardado en: {output_file}")

def CRPP_Hasta_4301a():
    # Definir el ancho fijo de las columnas crpp4601
    column_widths = [20, 31, 18, 9, 14]
    column_names = ["N°CUENTA", "PRESTATARIO", "SALDO", "TASA_EA", "SALDOxTASA"]

    # Ruta del directorio que contiene los archivos de texto
    directorio_in = "D:/CRPP/in"
    directorio_out = "D:/CRPP/out"

    # Lista para almacenar los DataFrames consolidados
    consolidated_data = []

    # Obtén la lista de archivos en el directorio
    archivos = os.listdir(directorio_in)

    for archivo in archivos:
        if archivo.lower().endswith(".txt"):
            ruta_completa = os.path.join(directorio_in, archivo)

            # Validar que el archivo contenga CRPP-4301 o crpp-4301
            with open(ruta_completa, 'r', encoding='utf-8') as file:
                contenido = file.read()
                if 'CRPP-4301' not in contenido and 'crpp-4301' not in contenido:
                    print(f"El archivo {archivo} no contiene CRPP-4301 o crpp-4301, será ignorado.")
                    continue

            # Leer línea por línea y extraer datos relevantes
            with open(ruta_completa, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()

            datos_encontrados = []
            dentro_de_seccion = False

            for i, line in enumerate(lines):
                if "hasta 360" in line:
                    dentro_de_seccion = True
                elif dentro_de_seccion and "mayor a 360" in line:
                    dentro_de_seccion = False

                if dentro_de_seccion:
                    if "Sub-total" in line:
                        # Retroceder y recoger las dos líneas anteriores
                        datos_relevantes = []
                        for j in range(i - 2, -1, -1):
                            if "------" in lines[j]:
                                break
                            datos_relevantes.append(lines[j].strip())
                        datos_encontrados.extend(reversed(datos_relevantes))
                    elif "Banco" in line:
                        # Manejar caso especial cuando no hay "Sub-total"
                        datos_relevantes = []
                        for j in range(i - 1, -1, -1):
                            if "------" in lines[j]:
                                break
                            datos_relevantes.append(lines[j].strip())
                        datos_encontrados.extend(reversed(datos_relevantes))

            # Convertir los datos encontrados a un DataFrame
            if datos_encontrados:
                data = []
                for dato in datos_encontrados:
                    partes = dato.split()
                    n_cuenta = partes[0]
                    prestatario = " ".join(partes[1:-3])
                    saldo = partes[-3].replace(",", "")
                    tasa_ea = partes[-2]
                    saldox_tasa = partes[-1].replace(",", "")
                    data.append([n_cuenta, prestatario, saldo, tasa_ea,saldox_tasa])

                filtered_data = pd.DataFrame(data, columns=column_names)
                #Convertir las columnas al tipo de dato adecuado
                filtered_data["SALDO"] = filtered_data["SALDO"].astype(float)
                #filtered_data["TASA_EA"] = filtered_data["TASA_EA"].astype(float)
                filtered_data["SALDOxTASA"] = filtered_data["SALDOxTASA"].astype(float)

                # Agregar los datos filtrados a las listas de datos consolidados
                consolidated_data.append(filtered_data)

    def guardar_excel(dataframes, output_path):
        # Concatenar todos los DataFrames consolidados
        if dataframes:
            resultado_df = pd.concat(dataframes, ignore_index=True)

            # Crear un nuevo Workbook con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Escribir el DataFrame al Worksheet
            for r_idx, row in enumerate(dataframe_to_rows(resultado_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Si es la primera fila (encabezados)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if c_idx in [3, 4, 5]:  # Columnas de montos
                            cell.number_format = '#,##0.00'

            # Ajustar el ancho de las columnas
            column_widths_adjusted = [25, 35, 20, 10, 15]
            for idx, width in enumerate(column_widths_adjusted, 1):
                ws.column_dimensions[chr(64 + idx)].width = width

            # Guardar el Workbook
            wb.save(output_path)
            print(f"El archivo consolidado se ha guardado en: {output_path}")
            # Mostrar data concatenada
            print(resultado_df)
        else:
            print("No se encontraron datos para consolidar.")

    # Guardar archivos Excel para datos con "MUNI" y otros datos
    output_file = os.path.join(directorio_out, "Reporte-Corto-Plazo-CRPP4301a-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    guardar_excel(consolidated_data, output_file)

def CRPP_Mayor_4301a():
    # Definir el ancho fijo de las columnas crpp4601
    column_widths = [20, 31, 18, 9, 14]
    column_names = ["N°CUENTA", "PRESTATARIO", "SALDO", "TASA_EA", "SALDOxTASA"]

    # Ruta del directorio que contiene los archivos de texto
    directorio_in = "D:/CRPP/in"
    directorio_out = "D:/CRPP/out"

    # Lista para almacenar los DataFrames consolidados
    consolidated_data = []

    # Obtén la lista de archivos en el directorio
    archivos = os.listdir(directorio_in)

    for archivo in archivos:
        if archivo.lower().endswith(".txt"):
            ruta_completa = os.path.join(directorio_in, archivo)

            # Validar que el archivo contenga CRPP-4301 o crpp-4301
            with open(ruta_completa, 'r', encoding='utf-8') as file:
                contenido = file.read()
                if 'CRPP-4301' not in contenido and 'crpp-4301' not in contenido:
                    print(f"El archivo {archivo} no contiene CRPP-4301 o crpp-4301, será ignorado.")
                    continue

            # Leer línea por línea y extraer datos relevantes
            with open(ruta_completa, 'r', encoding='iso-8859-1') as file:
                lines = file.readlines()

            datos_encontrados = []
            dentro_de_seccion = False

            for i, line in enumerate(lines):
                if "mayor a 360" in line:
                    dentro_de_seccion = True
                elif dentro_de_seccion and "hasta 360" in line:
                    dentro_de_seccion = False

                if dentro_de_seccion:
                    if "Sub-total" in line:
                        # Retroceder y recoger las dos líneas anteriores
                        datos_relevantes = []
                        for j in range(i - 2, -1, -1):
                            if "------" in lines[j]:
                                break
                            datos_relevantes.append(lines[j].strip())
                        datos_encontrados.extend(reversed(datos_relevantes))
                    elif "Banco" in line:
                        # Manejar caso especial cuando no hay "Sub-total"
                        datos_relevantes = []
                        for j in range(i - 1, -1, -1):
                            if "------" in lines[j]:
                                break
                            datos_relevantes.append(lines[j].strip())
                        datos_encontrados.extend(reversed(datos_relevantes))

            # Convertir los datos encontrados a un DataFrame
            if datos_encontrados:
                data = []
                for dato in datos_encontrados:
                    partes = dato.split()
                    n_cuenta = partes[0]
                    prestatario = " ".join(partes[1:-3])
                    saldo = partes[-3].replace(",", "")
                    tasa_ea = partes[-2]
                    saldox_tasa = partes[-1].replace(",", "")
                    data.append([n_cuenta, prestatario, saldo, tasa_ea,saldox_tasa])

                filtered_data = pd.DataFrame(data, columns=column_names)
                #Convertir las columnas al tipo de dato adecuado
                filtered_data["SALDO"] = filtered_data["SALDO"].astype(float)
                #filtered_data["TASA_EA"] = filtered_data["TASA_EA"].astype(float)
                filtered_data["SALDOxTASA"] = filtered_data["SALDOxTASA"].astype(float)

                # Agregar los datos filtrados a las listas de datos consolidados
                consolidated_data.append(filtered_data)

    def guardar_excel(dataframes, output_path):
        # Concatenar todos los DataFrames consolidados
        if dataframes:
            resultado_df = pd.concat(dataframes, ignore_index=True)

            # Crear un nuevo Workbook con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"

            # Escribir el DataFrame al Worksheet
            for r_idx, row in enumerate(dataframe_to_rows(resultado_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:  # Si es la primera fila (encabezados)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        if c_idx in [3, 4, 5]:  # Columnas de montos
                            cell.number_format = '#,##0.00'

            # Ajustar el ancho de las columnas
            column_widths_adjusted = [25, 35, 20, 10, 15]
            for idx, width in enumerate(column_widths_adjusted, 1):
                ws.column_dimensions[chr(64 + idx)].width = width

            # Guardar el Workbook
            wb.save(output_path)
            print(f"El archivo consolidado se ha guardado en: {output_path}")
            # Mostrar data concatenada
            print(resultado_df)
        else:
            print("No se encontraron datos para consolidar.")

    # Guardar archivos Excel para datos con "MUNI" y otros datos
    output_file = os.path.join(directorio_out, "Reporte-Largo-Plazo-CRPP4301a-" + datetime.datetime.now().strftime("%d-%m-%Y") + ".xlsx")
    guardar_excel(consolidated_data, output_file)

class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        self.title("BN_Operaciones")
        self.geometry("1000x562")

        self.resizable(False, False)

        # set grid layout 1x2
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # load images
        self.load_images()

        # set light mode as default
        customtkinter.set_appearance_mode("light")

        # initialize directories
        self.directorio_in = ""
        self.directorio_out = ""

        # create navigation frame
        self.navigation_frame = NavigationFrame(self)
        self.navigation_frame.grid(row=0, column=0, sticky="nsew")

        # create frames
        self.file_frame=FileFrame(self)
        self.home_frame = HomeFrame(self)
        self.second_frame = SecondFrame(self)
        self.third_frame = ThirdFrame(self)

        # select default frame
        self.select_frame_by_name("file")

    def load_images(self):
        image_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "img")
        self.logo_image = customtkinter.CTkImage(Image.open(os.path.join(image_path, "logo.png")), size=(150, 50))
        self.files = customtkinter.CTkImage(Image.open(os.path.join(image_path, "files.png")), size=(40, 40))
        self.cargar = customtkinter.CTkImage(Image.open(os.path.join(image_path, "corto_plazo.png")), size=(40, 40))
        self.seleccionar = customtkinter.CTkImage(Image.open(os.path.join(image_path, "largo_plazo.png")), size=(40, 40))
        self.descarga = customtkinter.CTkImage(Image.open(os.path.join(image_path, "Descargar.png")), size=(40, 40))

    def select_frame_by_name(self, name):
        # set button color for selected button
        self.navigation_frame.update_buttons(name)

        # show selected frame
        if name == "file":
            self.file_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.file_frame.grid_forget()
        if name == "home":
            self.home_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.home_frame.grid_forget()
        if name == "frame_2":
            self.second_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.second_frame.grid_forget()
        if name == "frame_3":
            self.third_frame.grid(row=0, column=1, sticky="nsew")
        else:
            self.third_frame.grid_forget()

class NavigationFrame(customtkinter.CTkFrame):
    def __init__(self, master):
        super().__init__(master, corner_radius=0)
        self.master = master

        self.grid_rowconfigure(5, weight=1)

        self.navigation_frame_label = customtkinter.CTkLabel(self, text="", image=self.master.logo_image,
                                                             compound="left", font=customtkinter.CTkFont(size=15, weight="bold"))
        self.navigation_frame_label.grid(row=0, column=0, padx=20, pady=20, sticky="s")

        self.file_button = customtkinter.CTkButton(self, corner_radius=0, height=40, border_spacing=10, text="Ruta de Archivos",font=customtkinter.CTkFont(size=12, weight="bold"),
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.master.files, anchor="w", command=lambda: self.master.select_frame_by_name("file"))
        self.file_button.grid(row=1, column=0, sticky="ew")

        self.home_button = customtkinter.CTkButton(self, corner_radius=0, height=40, border_spacing=10, text="R. Corto Plazo",font=customtkinter.CTkFont(size=12, weight="bold"),
                                                   fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                   image=self.master.cargar, anchor="w", command=lambda: self.master.select_frame_by_name("home"))
        self.home_button.grid(row=2, column=0, sticky="ew")

        self.frame_2_button = customtkinter.CTkButton(self, corner_radius=0, height=40, border_spacing=10, text="R. Largo Plazo",font=customtkinter.CTkFont(size=12, weight="bold"),
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.master.seleccionar, anchor="w", command=lambda: self.master.select_frame_by_name("frame_2"))
        self.frame_2_button.grid(row=3, column=0, sticky="ew")

        self.frame_3_button = customtkinter.CTkButton(self, corner_radius=0, height=40, border_spacing=10, text="CRPP 4301a",font=customtkinter.CTkFont(size=12, weight="bold"),
                                                      fg_color="transparent", text_color=("gray10", "gray90"), hover_color=("gray70", "gray30"),
                                                      image=self.master.descarga, anchor="w", command=lambda: self.master.select_frame_by_name("frame_3"))
        self.frame_3_button.grid(row=4, column=0, sticky="ew")

       

    def update_buttons(self, selected_frame):
        self.file_button.configure(fg_color="#FFFFFF" if selected_frame == "file" else "transparent")
        self.home_button.configure(fg_color="#FFFFFF" if selected_frame == "home" else "transparent")
        self.frame_2_button.configure(fg_color="#FFFFFF" if selected_frame == "frame_2" else "transparent")
        self.frame_3_button.configure(fg_color="#FFFFFF" if selected_frame == "frame_3" else "transparent")


class BaseFrame(customtkinter.CTkFrame):
    def __init__(self, master):
        super().__init__(master, corner_radius=0, fg_color="#FFFFFF")
        self.grid_columnconfigure(0, weight=1)

class FileFrame(BaseFrame):
    def __init__(self, master):
        super().__init__(master)

        self.file_title_label= customtkinter.CTkLabel(self, text="Definir ruta de carga y descarga de archivos",font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#BF0615")
        self.file_title_label.grid(row=0, column=0, padx=20, pady=30, sticky="nsew")

        #Frame de Ruta de los .txt
        self.file_upload_frame=customtkinter.CTkFrame(self, border_color="#B8B2B2", border_width=1, fg_color="transparent")
        self.file_upload_frame.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

        self.File_Entry_up = customtkinter.CTkEntry(master= self.file_upload_frame, placeholder_text="Ruta de la carpeta", width=400)
        self.File_Entry_up.grid(row=1, column=0, columnspan=6, padx=40, pady=20, sticky="n")

        self.File_Button_up= customtkinter.CTkButton(master=self.file_upload_frame, text="Seleccionar Carpeta", font=customtkinter.CTkFont(size=15,weight="bold"), 
                                                  border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0",
                                                  command=self.UpFile)
        self.File_Button_up.grid(row=1, column=8, padx=63, pady=20, sticky="n")

        #Frame de Ruta de descarga de los .xlsx
        self.file_download_frame=customtkinter.CTkFrame(self, border_color="#B8B2B2", border_width=1, fg_color="transparent")
        self.file_download_frame.grid(row=2, column=0, padx=20, pady=20, sticky="nsew")

        self.File_Entry_down = customtkinter.CTkEntry(master= self.file_download_frame, placeholder_text="Ruta de la carpeta", width=400)
        self.File_Entry_down.grid(row=1, column=0, columnspan=6, padx=40, pady=20, sticky="n")

        self.File_Button_down= customtkinter.CTkButton(master=self.file_download_frame, text="Seleccionar Carpeta", font=customtkinter.CTkFont(size=15,weight="bold"), 
                                                  border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0",
                                                  command=self.DowFile)
        self.File_Button_down.grid(row=1, column=8, padx=63, pady=20, sticky="n")

    def UpFile(self):
        file_path = filedialog.askdirectory()
        if file_path:
            self.File_Entry_up.delete(0, tkinter.END)
            self.File_Entry_up.insert(0, file_path)
            self.master.directorio_in = file_path  # Save the selected path

    def DowFile(self):
        file_path = filedialog.askdirectory()
        if file_path:
            self.File_Entry_down.delete(0, tkinter.END)
            self.File_Entry_down.insert(0, file_path)
            self.master.directorio_out = file_path  # Save the selected path


class HomeFrame(BaseFrame):
    def __init__(self, master):
        super().__init__(master)

        self.CRPP01_frame_large_image_label = customtkinter.CTkLabel(self, text="Reporte a Corto Plazo de CRPP 4601 - 4610",font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#BF0615")
        self.CRPP01_frame_large_image_label.grid(row=0, column=0, padx=20, pady=30, sticky="nsew")

        #Primer Frame
        self.Search_Frame=customtkinter.CTkFrame(self, border_color="#BF0615", border_width=1, fg_color="transparent")
        self.Search_Frame.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

        self.Search_Label=customtkinter.CTkLabel(master=self.Search_Frame, text="Ingresar palabra que desea extraer: ", font=customtkinter.CTkFont(size=15,weight="bold"))
        self.Search_Label.grid(row=0, column=0, padx=10, pady=10, sticky="")

        self.Search_Entry = customtkinter.CTkEntry(master= self.Search_Frame, placeholder_text="Buscar palabra: MUN, ELEC, EMP", width=400)
        self.Search_Entry.grid(row=1, column=0, columnspan=6, padx=40, pady=20, sticky="n")

        self.Search_Button= customtkinter.CTkButton(master=self.Search_Frame, text="Buscar", font=customtkinter.CTkFont(size=15,weight="bold"), border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0")
        self.Search_Button.grid(row=1, column=8, padx=63, pady=20, sticky="n")


        #Segundo Frame
        self.button_frame= customtkinter.CTkFrame(self, border_color="#BF0615", border_width=1, fg_color="transparent")
        self.button_frame.grid(row=2, column=0, padx=20, pady=20, sticky="nsew")

        self.radio_var = tkinter.IntVar(value = 0)

        self.label_radio_group = customtkinter.CTkLabel(master=self.button_frame, text="Seleccionar tipo de dato que desea extraer: ", font=customtkinter.CTkFont(size=15,weight="bold"))
        self.label_radio_group.grid(row=0, column=0, columnspan=2, padx=20, pady=10, sticky="")
        
        #Radio button para extaer los MUN
        self.mun_button= customtkinter.CTkRadioButton(master=self.button_frame,text="Municipales", font=customtkinter.CTkFont(size=15), variable=self.radio_var, value=0, hover_color="#F2CDD0", border_color="#7E7B7B", fg_color="#BF0615")
        self.mun_button.grid(row=1, column=0, pady=10, padx=30, sticky="n")

        self.otros_button= customtkinter.CTkRadioButton(master=self.button_frame,text="NO Municipales", font=customtkinter.CTkFont(size=15), variable=self.radio_var, value=1, hover_color="#F2CDD0", border_color="#7E7B7B", fg_color="#BF0615")
        self.otros_button.grid(row=1, column=2, pady=10, padx=20, sticky="n")

        self.cPlazo_button_ejecutar=customtkinter.CTkButton(master=self.button_frame, text="Ejecutar", font=customtkinter.CTkFont(size=15,weight="bold"), border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0", command = self.Ejecutar_CortoPlazo)
        self.cPlazo_button_ejecutar.grid(row=2, column=3, padx=20, pady=10, sticky="n")

    def Ejecutar_CortoPlazo(self):
        try:
            if hasattr(self.master, 'directorio_in') and hasattr(self.master, 'directorio_out'):
                Corto_Plazo_Mun(self.master.directorio_in, self.master.directorio_out)

                if self.radio_var.get() == 0:
                    Corto_Plazo_Mun()
                    CTkMessagebox(message="Datos extraidos exitosamente.",
                    icon="check", option_1="Ok", bg_color="#FFFFFF", fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
                elif self.radio_var.get() == 1:
                    Corto_plazo_No_Mun()
                    CTkMessagebox(message="Datos extraidos exitosamente.",
                    icon="check", option_1="Ok", bg_color="#FFFFFF", fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
                    
            else: 
                CTkMessagebox(title="Error", message="Por favor, selecciona las rutas de entrada y salida primero.", icon="cancel", option_1="OK")

        except Exception as e:
            CTkMessagebox(message=f"Ocurrió un error: {str(e)}",icon="cancel", option_1="Thanks", bg_color="#FFFFFF", 
                          fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
            
        finally:
            # Reiniciar el estado del radio button después de la ejecución
            self.radio_var.set(0)
            
class SecondFrame(BaseFrame):
    def __init__(self, master):
        super().__init__(master)

        self.filter_frame=customtkinter.CTkLabel(self, text="Reporte a Largo Plazo de CRPP 4602 - 4628",font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#BF0615")
        self.filter_frame.grid(row=0, column=0, padx=20, pady=30, sticky="nsew")

        #Primer Frame
        self.Search_Frame=customtkinter.CTkFrame(self, border_color="#BF0615", border_width=1, fg_color="transparent")
        self.Search_Frame.grid(row=1, column=0, padx=20, pady=20, sticky="nsew")

        self.Search_Label=customtkinter.CTkLabel(master=self.Search_Frame, text="Ingresar palabra que desea extraer: ", font=customtkinter.CTkFont(size=15,weight="bold"))
        self.Search_Label.grid(row=0, column=0, padx=10, pady=10, sticky="")

        self.Search_Entry = customtkinter.CTkEntry(master= self.Search_Frame, placeholder_text="Buscar palabra: MUN, ELEC, EMP", width=400)
        self.Search_Entry.grid(row=1, column=0, columnspan=6, padx=40, pady=20, sticky="n")

        self.Search_Button= customtkinter.CTkButton(master=self.Search_Frame, text="Buscar", font=customtkinter.CTkFont(size=15,weight="bold"), border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0")
        self.Search_Button.grid(row=1, column=8, padx=63, pady=20, sticky="n")

        #Segundo Frame
        self.button_frame= customtkinter.CTkFrame(self, border_color="#BF0615", border_width=1, fg_color="transparent")
        self.button_frame.grid(row=2, column=0, padx=20, pady=20, sticky="nsew")

        self.radio_var = tkinter.IntVar(value = 0)

        self.label_radio_group = customtkinter.CTkLabel(master=self.button_frame, text="Seleccionar tipo de dato que desea extraer: ", font=customtkinter.CTkFont(size=15,weight="bold"))
        self.label_radio_group.grid(row=0, column=0, columnspan=2, padx=20, pady=10, sticky="")
        
        #Radio button para extaer los MUN
        self.mun_button= customtkinter.CTkRadioButton(master=self.button_frame,text="Municipales", font=customtkinter.CTkFont(size=15), variable=self.radio_var, value=0, hover_color="#F2CDD0", border_color="#7E7B7B", fg_color="#BF0615")
        self.mun_button.grid(row=1, column=0, pady=10, padx=30, sticky="n")

        self.otros_button= customtkinter.CTkRadioButton(master=self.button_frame,text="NO Municipales", font=customtkinter.CTkFont(size=15), variable=self.radio_var, value=1, hover_color="#F2CDD0", border_color="#7E7B7B", fg_color="#BF0615")
        self.otros_button.grid(row=1, column=2, pady=10, padx=20, sticky="n")

        self.cPlazo_button_ejecutar=customtkinter.CTkButton(master=self.button_frame, text="Ejecutar", font=customtkinter.CTkFont(size=15,weight="bold"), border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0", command=self.Ejecutar_LargoPlazo)
        self.cPlazo_button_ejecutar.grid(row=2, column=3, padx=20, pady=10, sticky="n")

    def Ejecutar_LargoPlazo(self):
        try:
            if self.radio_var.get() == 0:
                Largo_Plazo_Mun()
                CTkMessagebox(message="Datos extraidos exitosamente.",
                  icon="check", option_1="Ok", bg_color="#FFFFFF", fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
            else:
                Largo_Plazo_No_Mun()
                CTkMessagebox(message="Datos extraidos exitosamente.",
                  icon="check", option_1="Ok", bg_color="#FFFFFF", fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")

        except Exception as e:
            CTkMessagebox(message=f"Ocurrió un error: {str(e)}",icon="cancel", option_1="Thanks", bg_color="#FFFFFF", 
                          fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
            
        finally:
            # Reiniciar el estado del radio button después de la ejecución
            self.radio_var.set(0)

class ThirdFrame(BaseFrame):
    def __init__(self, master):
        super().__init__(master)

        self.CRPP4301_frame=customtkinter.CTkLabel(self, text="Reporte CRPP 4301a",font=customtkinter.CTkFont(size=20, weight="bold"), text_color="#BF0615")
        self.CRPP4301_frame.grid(row=0, column=0, padx=20, pady=30, sticky="nsew")

        self.CRPP4301_Button=customtkinter.CTkButton(self, text="Ejecutar", font=customtkinter.CTkFont(size=15,weight="bold"), border_width=1, border_color="#BF0615", fg_color="#FFFFFF", text_color="#BF0615", hover_color="#F2CDD0", command=self.Ejecutar_CRPP4301a)
        self.CRPP4301_Button.grid(row=1, column=0, padx=20, pady=10, sticky="n")

    def Ejecutar_CRPP4301a(self):
        try:
            CRPP_Hasta_4301a()
            CRPP_Mayor_4301a()
            CTkMessagebox(message="Datos extraidos exitosamente.",
                  icon="check", option_1="Ok", bg_color="#FFFFFF", fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")
        except Exception as e:
            CTkMessagebox(message=f"Ocurrió un error: {str(e)}",icon="cancel", option_1="Thanks", bg_color="#FFFFFF", 
                          fg_color="#FFFFFF", button_color="#4D4D4D", button_hover_color="#BF0615", button_text_color="#FFFFFF")

if __name__ == "__main__":
    app = App()
    app.mainloop()
